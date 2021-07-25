from ciscoconfparse import CiscoConfParse
import pandas
import os
import openpyxl
from openpyxl.styles import Font
import string
  
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent

def main(file_name, checklistOptions):
    # print("print inside interface", checklistOptions)
    file_to_parse = CiscoConfParse(file_name)
    empty_dic = {}
    for each_file_to_parse in file_to_parse.find_objects('^interface'):
        empty_dic[each_file_to_parse.text] = []

    for each_interface in empty_dic:
        interface_children_commands = file_to_parse.find_all_children(each_interface+"$")

        if "DESCRIPTION" in checklistOptions:
            if any("description" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                description_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "description" in each_interface_children_commands)
                empty_dic[each_interface].append(description_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "IP ADDRESS" in checklistOptions:
            if any("ip address" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                ip_address_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "ip address" in each_interface_children_commands)
                empty_dic[each_interface].append(ip_address_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT MODE" in checklistOptions:
            if any("switchport mode" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_mode_trunk_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport mode" in each_interface_children_commands)
                empty_dic[each_interface].append(switchport_mode_trunk_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT ACCESS VLAN" in checklistOptions:
            if any("switchport access vlan" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_access_vlan_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport access vlan" in each_interface_children_commands)
                empty_dic[each_interface].append(switchport_access_vlan_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT VOICE VLAN" in checklistOptions:
            if any("switchport voice vlan" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_voice_vlan_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport voice vlan" in each_interface_children_commands)
                empty_dic[each_interface].append(switchport_voice_vlan_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT TRUNK NATIVE" in checklistOptions:
            if any("switchport trunk native" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_trunk_native_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport trunk native" in each_interface_children_commands)
                empty_dic[each_interface].append(switchport_trunk_native_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT TRUNK ALLOWED VLAN" in checklistOptions:
            if any("switchport trunk allowed vlan" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_trunk_allowed_vlan_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport trunk allowed vlan" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_switchport_trunk_allowed_vlan_children for each_switchport_trunk_allowed_vlan_children in switchport_trunk_allowed_vlan_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "IP HELPER-ADDRESS" in checklistOptions:
            if any("ip helper-address" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                    ip_helper_address_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "ip helper-address" in each_interface_children_commands)
                    empty_dic[each_interface].append(str([each_ip_helper_address_children for each_ip_helper_address_children in ip_helper_address_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT NONEGOTIATE" in checklistOptions:
            if any("switchport nonegotiate" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_nonegotiate_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport nonegotiate" in each_interface_children_commands)
                empty_dic[each_interface].append(switchport_nonegotiate_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "LOGGING" in checklistOptions:
            if any("logging" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                logging_event_trunk_status_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "logging" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_logging_event_trunk_status_children for each_logging_event_trunk_status_children in logging_event_trunk_status_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "UDLD PORT AGGRESSIVE" in checklistOptions:
            if any("udld port aggressive" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                udld_port_aggressive_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "udld port aggressive" in each_interface_children_commands)
                empty_dic[each_interface].append(udld_port_aggressive_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "IP DHCP SNOOPING" in checklistOptions:
            if any("ip dhcp snooping" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                ip_dhcp_snooping_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "ip dhcp snooping" in each_interface_children_commands)
                empty_dic[each_interface].append(ip_dhcp_snooping_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SPANNING-TREE LINK TYPE" in checklistOptions:
            if any("spanning-tree link-type" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                spanning_tree_link_type_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "spanning-tree link-type" in each_interface_children_commands)
                empty_dic[each_interface].append(spanning_tree_link_type_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SWITCHPORT PORT-SECURITY" in checklistOptions:
            if any("switchport port-security" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                switchport_port_security_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "switchport port-security" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_switchport_port_security_children for each_switchport_port_security_children in switchport_port_security_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "DEVICE TRACKING" in checklistOptions:
            if any("device" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                ip_device_tracking_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "device" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_ip_device_tracking_children for each_ip_device_tracking_children in ip_device_tracking_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SRR-QUEUE" in checklistOptions:
            if any("srr-queue" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                srr_queue_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "srr-queue" in each_interface_children_commands)
                empty_dic[each_interface].append(srr_queue_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SPANNING-TREE PORTFAST" in checklistOptions:
            if any("spanning-tree portfast" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                spanning_tree_portfast_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "spanning-tree portfast" in each_interface_children_commands)
                empty_dic[each_interface].append(spanning_tree_portfast_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SPANNING-TREE BPDUGUARD" in checklistOptions:
            if any("spanning-tree bpduguard" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                spanning_tree_bpduguard_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "spanning-tree bpduguard" in each_interface_children_commands)
                empty_dic[each_interface].append(spanning_tree_bpduguard_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SPANNING-TREE GUARD ROOT" in checklistOptions:
            if any("spanning-tree guard root" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                spanning_tree_guard_root_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "spanning-tree guard root" in each_interface_children_commands)
                empty_dic[each_interface].append(spanning_tree_guard_root_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SERVICE POLICY" in checklistOptions:
            if any("service-policy" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                service_policy_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "service-policy" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_service_policy_children for each_service_policy_children in service_policy_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "PIM MODE" in checklistOptions:
            if any("ip pim" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                ip_pim_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "ip pim" in each_interface_children_commands)
                empty_dic[each_interface].append(ip_pim_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "REDIRECTS" in checklistOptions:
            if any("redirects" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                redirects_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "redirects" in each_interface_children_commands)
                empty_dic[each_interface].append(redirects_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "UNREACHABLES" in checklistOptions:
            if any("unreachables" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                unreachables_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "unreachables" in each_interface_children_commands)
                empty_dic[each_interface].append(unreachables_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "ACCESS-GROUP" in checklistOptions:
            if any("access-group" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                access_group_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "access-group" in each_interface_children_commands)
                empty_dic[each_interface].append(str([each_access_group_children for each_access_group_children in access_group_children]).replace("[","").replace("]","").replace("'","")[1:])
            else:
                empty_dic[each_interface].append("N/A")
        if "SHUTDOWN" in checklistOptions:
            if any("shutdown" in each_interface_children_commands for each_interface_children_commands in interface_children_commands):
                shutdown_children = list(each_interface_children_commands for each_interface_children_commands in interface_children_commands if "shutdown" in each_interface_children_commands)
                empty_dic[each_interface].append(shutdown_children[0][1:])
            else:
                empty_dic[each_interface].append("N/A")
    # print(os.path.join(BASE_DIR,'TEMP_FILE_STORAGE/interface_testing.xlsx'))
    filePath = os.path.join(BASE_DIR,'TEMP_FILE_STORAGE/interface_testing.xlsx')
    wb = openpyxl.load_workbook(filePath)
    wb.create_sheet("interface", -1)
    wb.active = 1
    sheet = wb.active
    row = 1
    conf_header = ['NTP','AAA','LOGS','CLOCK INFORMATION','SERVICES','SNMP']
    finalList = []
    for header in checklistOptions:
        if header not in conf_header:
            finalList.append(header)
    # print(finalList)
    for count,head in enumerate(finalList):
        count = count + 1
        sheet.cell(row = row, column = count).value = head
        sheet.cell(row = row, column = count).font = Font(bold=True)
        # print(head, count)
    row = row + 1 
    for key, items in empty_dic.items():
        sheet.cell(row = row, column=1).value = key
        for index, item in enumerate(items):
            sheet.cell(row = row, column=index+2).value = item
            # print(index)
        row = row + 1 
            # sheet.cell(row = xbar, column = count).value = items[key]
    
    for letter in string.ascii_uppercase[:len(checklistOptions)]:
        sheet.column_dimensions[letter].width = 25
    wb.save(filePath)

    # with open('/home/ec2-user/webpage/Cisco_Parser/TEMP_FILE_STORAGE/interface_testing.csv', 'a', newline='') as file:
    #     writer = csv.writer(file)
    #     writer.writerow(["\n"])
    #     writer.writerow(["INTERFACE", "DESCRIPTION", "IP ADDRESS", "SWITCHPORT MODE", "SWITCHPORT ACCESS VLAN", "SWITCHPORT VOICE VLAN", "SWITCHPORT TRUNK NATIVE","SWITCHPORT TRUNK ALLOWED VLAN", "IP HELPER-ADDRESS", "SWITCHPORT NONEGOTIATE", "LOGGING", "UDLD PORT AGGRESSIVE", "IP DHCP SNOOPING", "SPANNING-TREE LINK TYPE","SWITCHPORT PORT-SECURITY", "DEVICE TRACKICKING", "SRR-QUEUE", "SPANNING-TREE PORTFAST", "SPANNING-TREE BPDUGUARD", "SPANNING-TREE GUARD ROOT", "SERVICE POLICY", "PIM MODE", "REDIRECTS", "UNREACHABLES", "ACCESS-GROUP", "SHUTDOWN"])
    #     for key, items in empty_dic.items():
    #         writer.writerow([key,items[0],items[1],items[2],items[3],items[4],items[5],items[6],items[7],items[8],items[9],items[10],items[11],items[12],items[13],items[14],items[15],items[16],items[17],items[18],items[19],items[20],items[21],items[22],items[23],items[24]])
    
    #reading_csv_pandas = pandas.read_csv("interface_testing.csv")
    #print(reading_csv_pandas.head())
    
if __name__ == "__main__":
    main()


